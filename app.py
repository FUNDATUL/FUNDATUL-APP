from pathlib import Path
from datetime import date, datetime
import shutil
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from supabase import create_client
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
try:
    prueba_supabase = supabase.table("personas").select("id").limit(1).execute()
    conexion_supabase_ok = True
except Exception as e:
    conexion_supabase_ok = False
    error_supabase = str(e)

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR 
SOURCE_XLSX = DATA_DIR / "sistema_fundatul.xlsx"
WORK_XLSX = DATA_DIR / "sistema_fundatul_trabajo.xlsx"

st.set_page_config(page_title="FUNDATUL · Vida Independiente", page_icon="🏠", layout="wide")
if conexion_supabase_ok:
    st.success("Supabase conectado correctamente")
else:
    st.error("No se pudo conectar con Supabase")
    st.code(error_supabase)           
st.markdown("""
<style>
.block-container {padding-top: 1.5rem; padding-bottom: 2rem;}
[data-testid="stMetric"] {background: #f7f7f8; border: 1px solid #e6e6e8; padding: 14px; border-radius: 12px;}
.small-note {font-size: .88rem; opacity: .78;}
</style>
""", unsafe_allow_html=True)

if not WORK_XLSX.exists():
    shutil.copy2(SOURCE_XLSX, WORK_XLSX)

TABLES = {
    "Personas": ("Configuracion", 4),
    "Intervenciones": ("Intervenciones", 4),
    "Incidencias": ("Incidencias", 4),
    "Objetivos PIA": ("Objetivos PIA", 4),
    "Coordinaciones": ("Coordinaciones", 4),
    "Seguimiento semanal": ("Seguimiento semanal", 4),
    "Seguimiento mensual": ("Seguimiento mensual", 4),
    "Indicadores": ("Indicadores", 4),
}

@st.cache_data(show_spinner=False)
def read_table(path_str, sheet, header_row, stamp):
    df = pd.read_excel(path_str, sheet_name=sheet, header=header_row-1, engine="openpyxl")
    df = df.dropna(how="all")
    # Remove template rows that only contain sequential IDs / formulas and no user data.
    non_id_cols = [c for c in df.columns if str(c).strip().lower() not in {"id", "mes"}]
    if non_id_cols:
        meaningful = df[non_id_cols].notna().any(axis=1)
        df = df[meaningful].copy()
    return df

def stamp(path):
    p = Path(path)
    return p.stat().st_mtime_ns if p.exists() else 0

def load_table(name):
    sheet, header = TABLES[name]
    return read_table(str(WORK_XLSX), sheet, header, stamp(WORK_XLSX))

def clear_cache():
    read_table.clear()

def excel_safe(v):
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if pd.isna(v):
        return None
    return v

def append_record(table_name, record):
    sheet, header_row = TABLES[table_name]
    wb = load_workbook(WORK_XLSX)
    ws = wb[sheet]
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    # Find first truly blank data row. Template IDs do not count as user data.
    row = header_row + 1
    while row <= ws.max_row:
        vals = [ws.cell(row, c).value for c in range(1, ws.max_column + 1)]
        # user-filled if any field other than ID/Mes has a value
        user_vals = []
        for h, v in zip(headers, vals):
            if str(h).strip().lower() not in {"id", "mes"}:
                user_vals.append(v)
        if not any(v not in (None, "") for v in user_vals):
            break
        row += 1
    if row > ws.max_row:
        ws.insert_rows(row)
    # Preserve/create sequential ID when available.
    if "ID" in headers and not record.get("ID"):
        ids = []
        id_col = headers.index("ID") + 1
        for r in range(header_row + 1, ws.max_row + 1):
            v = ws.cell(r, id_col).value
            if isinstance(v, (int, float)):
                ids.append(int(v))
        record["ID"] = max(ids, default=0) + 1
    for c, h in enumerate(headers, start=1):
        if h in record:
            ws.cell(row, c).value = excel_safe(record[h])
        elif h == "Mes" and record.get("Fecha"):
            f = record["Fecha"]
            if isinstance(f, (date, datetime)):
                ws.cell(row, c).value = f.strftime("%Y-%m")
    wb.save(WORK_XLSX)
    clear_cache()

def active_people():
    df = load_table("Personas")
    if df.empty:
        return []
    name_col = "Nombre / código" if "Nombre / código" in df.columns else df.columns[1]
    if "Estado" in df.columns:
        df = df[df["Estado"].fillna("").astype(str).str.lower().eq("activa")]
    return [str(x) for x in df[name_col].dropna().tolist()]

def person_id_for_name(name):
    df = load_table("Personas")
    if df.empty: return name
    m = df[df["Nombre / código"].astype(str).eq(str(name))]
    return str(m.iloc[0]["ID persona"]) if not m.empty else name

def person_filter(df, person):
    if df.empty: return df
    candidates = [c for c in ["Persona", "Persona / ámbito", "Nombre / código"] if c in df.columns]
    if not candidates: return df
    pid = person_id_for_name(person)
    c = candidates[0]
    return df[df[c].astype(str).isin([str(person), str(pid)])]

def show_table(df, key):
    if df.empty:
        st.info("Todavía no hay registros.")
        return
    st.dataframe(df, use_container_width=True, hide_index=True, key=key)

with st.sidebar:
    st.title("FUNDATUL")
    st.caption("Servicio de Apoyo a la Vida Independiente")
    section = st.radio("Ir a", ["Dashboard", "Personas", "PIA inicial", "Intervenciones", "Incidencias", "Objetivos PIA", "Coordinaciones", "Datos / copia"])
    st.divider()
    st.caption("Prototipo basado en el Excel original. Los cambios se guardan en una copia de trabajo del archivo.")

st.title("Sistema de Gestión y Seguimiento")
st.caption("Prototipo web · Vida Independiente")

if section == "Dashboard":
    personas = load_table("Personas")
    interv = load_table("Intervenciones")
    incid = load_table("Incidencias")
    obj = load_table("Objetivos PIA")
    coord = load_table("Coordinaciones")

    active = 0
    if not personas.empty and "Estado" in personas:
        active = int(personas["Estado"].fillna("").astype(str).str.lower().eq("activa").sum())
    hours = float(pd.to_numeric(interv.get("Duración (h)", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    n_interv = len(interv)
    n_incid = len(incid)
    n_obj_active = int(obj.get("Estado", pd.Series(dtype=str)).fillna("").astype(str).str.lower().eq("activo").sum())
    n_obj_done = int(obj.get("Estado", pd.Series(dtype=str)).fillna("").astype(str).str.lower().eq("alcanzado").sum())
    grav = incid.get("Gravedad", pd.Series(dtype=str)).fillna("").astype(str).str.lower()
    n_graves = int(grav.isin(["grave", "muy grave"]).sum())

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Personas activas", active)
    c2.metric("Horas de apoyo", f"{hours:.1f}")
    c3.metric("Intervenciones", n_interv)
    c4.metric("Incidencias", n_incid)
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Objetivos activos", n_obj_active)
    c2.metric("Objetivos alcanzados", n_obj_done)
    c3.metric("Incidencias graves", n_graves)
    c4.metric("Coordinaciones", len(coord))

    st.subheader("Actividad reciente")
    if not interv.empty:
        recent = interv.copy()
        if "Fecha" in recent.columns:
            recent["Fecha"] = pd.to_datetime(recent["Fecha"], errors="coerce")
            recent = recent.sort_values("Fecha", ascending=False)
        show_table(recent.head(10), "dash_recent")
    else:
        st.info("Registra una intervención para empezar a alimentar el cuadro de mando.")

elif section == "Personas":
    respuesta_personas = supabase.table("personas").select("*").execute()
    personas = pd.DataFrame(respuesta_personas.data)
    st.subheader("Personas")
    show_table(personas, "people")
    people = personas["nombre"].dropna().astype(str).tolist()
    if people:
        st.subheader("Ficha individual")
        p = st.selectbox("Seleccionar persona", people)
        interv = person_filter(load_table("Intervenciones"), p)
        incid = person_filter(load_table("Incidencias"), p)
        obj = person_filter(load_table("Objetivos PIA"), p)
        coord = person_filter(load_table("Coordinaciones"), p)
        hrs = pd.to_numeric(interv.get("Duración (h)", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
        avance = pd.to_numeric(obj.get("Grado consecución %", pd.Series(dtype=float)), errors="coerce").dropna()
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("Horas de apoyo", f"{hrs:.1f}")
        c2.metric("Intervenciones", len(interv))
        c3.metric("Incidencias", len(incid))
        c4.metric("Avance medio", f"{avance.mean():.0f}%" if len(avance) else "—")
        tabs = st.tabs(["Objetivos PIA", "Intervenciones", "Incidencias", "Coordinaciones"])
        with tabs[0]: show_table(obj, "p_obj")
        with tabs[1]: show_table(interv, "p_int")
        with tabs[2]: show_table(incid, "p_inc")
        with tabs[3]: show_table(coord, "p_coord")
elif section == "PIA inicial":
    st.subheader("PIA inicial")
    st.caption("Plan Individual de Apoyos · FUNDATUL · DEJAR SER")

    respuesta_personas = supabase.table("personas").select("*").execute()
    personas = pd.DataFrame(respuesta_personas.data)
    nombres_personas = personas["nombre"].dropna().astype(str).tolist()

    if not nombres_personas:
        st.warning("No hay personas registradas.")
    else:
        persona_sel = st.selectbox(
            "Selecciona una persona",
            nombres_personas,
            key="pia_persona"
        )
        persona_data = (
            supabase.table("personas")
            .select("id")
            .eq("nombre", persona_sel)
            .limit(1)
            .execute()
        )

        persona_id = persona_data.data[0]["id"] if persona_data.data else None

        pia_existente = None
        valoraciones_existentes = []
        objetivos_existentes = []

        if persona_id:
            respuesta_pia = (
                supabase.table("pia_inicial")
                .select("*")
                .eq("persona_id", persona_id)
                .order("id", desc=True)
                .limit(1)
                .execute()
            )

            if respuesta_pia.data:
                pia_existente = respuesta_pia.data[0]
                pia_id_existente = pia_existente["id"]

                valoraciones_existentes = (
                    supabase.table("valoracion_funcional")
                    .select("*")
                    .eq("pia_id", pia_id_existente)
                    .execute()
                ).data or []

                objetivos_existentes = (
                    supabase.table("objetivos_pia")
                    .select("*")
                    .eq("pia_id", pia_id_existente)
                    .execute()
                ).data or []
        st.markdown("### 0. Control documental e identificación")
        codigo = st.text_input(
    "Código de participante",
    value=(pia_existente.get("codigo_participante", "") if pia_existente else "")
)
        fecha_nacimiento = st.date_input(
    "Fecha de nacimiento",
    value=(
        pd.to_datetime(pia_existente.get("fecha_nacimiento")).date()
        if pia_existente and pia_existente.get("fecha_nacimiento")
        else None
    )
)
        municipio = st.text_input(
    "Municipio / entorno habitual",
    value=(pia_existente.get("municipio", "") if pia_existente else "")
)

profesional = st.text_input(
    "Profesional referente",
    value=(pia_existente.get("profesional_referente", "") if pia_existente else "")
)

col1, col2 = st.columns(2)
    with col1:
            periodo_desde = st.date_input(
    "Periodo del PIA · Desde",
    value=(
        pd.to_datetime(pia_existente.get("periodo_desde")).date()
        if pia_existente and pia_existente.get("periodo_desde")
        else None
    )
)
    with col2:
            periodo_hasta = st.date_input(
    "Periodo del PIA · Hasta",
    value=(
        pd.to_datetime(pia_existente.get("periodo_hasta")).date()
        if pia_existente and pia_existente.get("periodo_hasta")
        else None
    )
)

opciones_version = ["Inicial", "Revisión intermedia", "Final", "Extraordinaria"]
version_guardada = pia_existente.get("version", "Inicial") if pia_existente else "Inicial"

version = st.selectbox(
    "Versión",
    opciones_version,
    index=opciones_version.index(version_guardada) if version_guardada in opciones_version else 0
)

 st.markdown("### 2. Perfil personal, preferencias y voz de la persona")
 situacion_actual = st.text_area(
    "2.1. Situación actual y contexto de vida",
    value=pia_existente.get("situacion_actual", "") if pia_existente else ""
)
 importante_para = st.text_area(
    "2.2. Lo que es importante PARA la persona",
    value=pia_existente.get("importante_para", "") if pia_existente else ""
)
 bienestar_seguridad = st.text_area(
    "2.3. Lo que es importante POR su bienestar y seguridad",
    value=pia_existente.get("importante_por", "") if pia_existente else ""
)
 fortalezas = st.text_area(
    "2.4. Capacidades, fortalezas, intereses y recursos personales",
    value=pia_existente.get("capacidades_fortalezas", "") if pia_existente else ""
)
 comunicacion = st.text_area(
    "2.5. Comunicación y apoyo a la toma de decisiones",
    value=pia_existente.get("comunicacion_toma_decisiones", "") if pia_existente else ""
)
 red_apoyo = st.text_area(
    "2.6. Red de apoyo y personas significativas",
    value=pia_existente.get("red_apoyo", "") if pia_existente else ""
)
 cambios_deseados = st.text_area(
    "2.7. Cambios que la persona quiere conseguir",
    value=pia_existente.get("cambios_deseados", "") if pia_existente else ""
)

 st.markdown("### 3. Valoración funcional inicial por áreas de vida")
        st.caption(
            "Escala 0–4: 0 = autónomo/a · 1 = supervisión puntual · "
            "2 = apoyo intermitente · 3 = apoyo frecuente · 4 = apoyo intenso/estable."
        )

        areas_vida = [
            "Autocuidado e higiene",
            "Organización doméstica y rutinas",
            "Alimentación y vida diaria",
            "Manejo funcional del dinero y compras",
            "Movilidad, transporte y orientación",
            "Gestiones y acceso a recursos/servicios",
            "Competencia digital y uso funcional de TIC",
            "Bienestar y regulación emocional",
            "Comunicación, relaciones y límites",
            "Participación comunitaria y ocio",
            "Empleo / formación / actividad ocupacional",
            "Toma de decisiones y autodeterminación",
        ]

        valoracion_funcional = []

        for i, area in enumerate(areas_vida):
            valoracion_existente = next(
                (v for v in valoraciones_existentes if v.get("area") == area),
                None
            )
            st.markdown(f"**{area}**")
            evidencia = st.text_area(
    "Situación inicial / evidencia",
    value=valoracion_existente.get("situacion_inicial_evidencia", "") if valoracion_existente else "",
    key=f"pia_evidencia_{i}"
)

            c1, c2, c3 = st.columns(3)

            with c1:
                nivel_guardado = valoracion_existente.get("nivel_apoyo", 0) if valoracion_existente else 0

                nivel = st.selectbox(
                    "Nivel de apoyo",
    [0, 1, 2, 3, 4],
    index=[0, 1, 2, 3, 4].index(nivel_guardado) if nivel_guardado in [0, 1, 2, 3, 4] else 0,
    key=f"pia_nivel_{i}"
)

            with c2:
                prioridad_guardada = valoracion_existente.get("prioridad", "Alta") if valoracion_existente else "Alta"
opciones_prioridad = ["Alta", "Media", "Baja"]

prioridad = st.selectbox(
    "Prioridad",
    opciones_prioridad,
    index=opciones_prioridad.index(prioridad_guardada) if prioridad_guardada in opciones_prioridad else 0,
    key=f"pia_prioridad_{i}"
)

            with c3:
    contexto_guardado = []
    if valoracion_existente and valoracion_existente.get("contexto_principal"):
        contexto_guardado = [
            x.strip() for x in valoracion_existente.get("contexto_principal", "").split(",")
            if x.strip()
    ]

    contexto = st.multiselect(
        "Contexto principal",
        ["Hogar", "Comunidad", "Sede", "Online"],
        default=contexto_guardado,
        key=f"pia_contexto_{i}"
    )

            valoracion_funcional.append({
                "Área": area,
                "Situación inicial / evidencia": evidencia,
                "Nivel apoyo": nivel,
                "Prioridad": prioridad,
                "Contexto": ", ".join(contexto),
            })

            st.divider()

        sintesis_prioridades = st.text_area(
            "Síntesis de prioridades acordadas"
        )
        st.markdown("### 4. Resultados personales y objetivos priorizados")
        st.caption(
            "Cada objetivo debe indicar qué cambio se espera, "
            "cómo se observará y con qué evidencia se considerará avanzado o alcanzado."
        )

        objetivos_pia_inicial = []

        for i in range(1, 7):
           objetivo_existente = next(
        (o for o in objetivos_existentes if o.get("numero_objetivo") == i),
        None
    )
            st.markdown(f"**Objetivo {i}**")

            resultado = st.text_area(
                "Resultado personal esperado",
                value=objetivo_existente.get("resultado_esperado", "") if objetivo_existente else "",
                key=f"pia_obj_resultado_{i}"
            )

            punto_partida = st.text_area(
                "Punto de partida",
                value=objetivo_existente.get("punto_partida", "") if objetivo_existente else "",
                key=f"pia_obj_partida_{i}"
            )

            indicador = st.text_area(
                "Indicador / evidencia",
                value=objetivo_existente.get("indicador_evidencia", "") if objetivo_existente else "",
                key=f"pia_obj_indicador_{i}"
            )

            c1, c2, c3 = st.columns(3)

            with c1:
                meta = st.text_input(
                    "Meta / criterio",
                    value=objetivo_existente.get("meta_criterio", "") if objetivo_existente else "",
                    key=f"pia_obj_meta_{i}"
                )

            with c2:
                fecha_revision_guardada = (
                    pd.to_datetime(objetivo_existente.get("fecha_revision")).date()
                    if objetivo_existente and objetivo_existente.get("fecha_revision")
                    else None
        )

               fecha_revision = st.date_input(
                   "Fecha de revisión",
                   value=fecha_revision_guardada,
                   key=f"pia_obj_fecha_{i}"
        )

            with c3:
                prioridad_guardada = objetivo_existente.get("prioridad", "Alta") if objetivo_existente else "Alta"
                opciones_prioridad_obj = ["Alta", "Media", "Baja"]

                prioridad_obj = st.selectbox(
                 "Prioridad",
                 opciones_prioridad_obj,
                 index=opciones_prioridad_obj.index(prioridad_guardada) if prioridad_guardada in opciones_prioridad_obj else 0,
                 key=f"pia_obj_prioridad_{i}"
)

            objetivos_pia_inicial.append({
                "Objetivo": i,
                "Resultado personal esperado": resultado,
                "Punto de partida": punto_partida,
                "Indicador / evidencia": indicador,
                "Meta / criterio": meta,
                "Fecha revisión": fecha_revision,
                "Prioridad": prioridad_obj,
            })

            st.divider()

        objetivo_general = st.text_area(
            "Objetivo general del periodo"
            value=pia_existente.get("objetivo_general", "") if pia_existente else ""
        )

        if st.button("Guardar PIA inicial", type="primary"):
            try:
                persona_data = (
                    supabase.table("personas")
                    .select("id")
                    .eq("nombre", persona_sel)
                    .limit(1)
                    .execute()
                )

                if not persona_data.data:
                    st.error("No se ha encontrado la persona seleccionada en Supabase.")
                else:
                    persona_id = persona_data.data[0]["id"]

                    pia_data = {
                        "persona_id": persona_id,
                        "fecha_nacimiento": fecha_nacimiento.isoformat() if fecha_nacimiento else None,
                        "municipio": municipio,
                        "profesional_referente": profesional,
                        "periodo_desde": periodo_desde.isoformat() if periodo_desde else None,
                        "periodo_hasta": periodo_hasta.isoformat() if periodo_hasta else None,
                        "version": version,
                        "situacion_actual": situacion_actual,
                        "importante_para": importante_para,
                        "importante_por": bienestar_seguridad,
                        "capacidades_fortalezas": fortalezas,
                        "comunicacion_toma_decisiones": comunicacion,
                        "red_apoyo": red_apoyo,
                        "cambios_deseados": cambios_deseados,
                        "sintesis_prioridades": sintesis_prioridades,
                        "objetivo_general": objetivo_general,
                        "estado": "Activo"
                    }
                        

                    if pia_existente:
    pia_id = pia_existente["id"]

    supabase.table("pia_inicial") \
        .update(pia_data) \
        .eq("id", pia_id) \
        .execute()

    supabase.table("valoracion_funcional") \
        .delete() \
        .eq("pia_id", pia_id) \
        .execute()

    supabase.table("objetivos_pia") \
        .delete() \
        .eq("pia_id", pia_id) \
        .execute()

else:
    pia_respuesta = (
        supabase.table("pia_inicial")
        .insert(pia_data)
        .execute()
    )

    pia_id = pia_respuesta.data[0]["id"]

for valoracion in valoracion_funcional:
    supabase.table("valoracion_funcional").insert({
        "pia_id": pia_id,
        "persona_id": persona_id,
        "area": valoracion["Área"],
        "situacion_inicial_evidencia": valoracion["Situación inicial / evidencia"],
        "nivel_apoyo": valoracion["Nivel apoyo"],
        "prioridad": valoracion["Prioridad"],
        "contexto_principal": valoracion["Contexto"]
    }).execute()

for objetivo in objetivos_pia_inicial:
    supabase.table("objetivos_pia").insert({
        "pia_id": pia_id,
        "persona_id": persona_id,
        "numero_objetivo": objetivo["Objetivo"],
        "resultado_esperado": objetivo["Resultado personal esperado"],
        "punto_partida": objetivo["Punto de partida"],
        "indicador_evidencia": objetivo["Indicador / evidencia"],
        "meta_criterio": objetivo["Meta / criterio"],
        "fecha_revision": objetivo["Fecha revisión"].isoformat() if objetivo["Fecha revisión"] else None,
        "prioridad": objetivo["Prioridad"],
        "estado": "Activo"
    }).execute()

    st.success("PIA inicial guardado correctamente en Supabase.")

            except Exception as e:
                st.error(f"No se pudo guardar el PIA: {e}")

elif section == "Intervenciones":
    st.subheader("Registrar intervención")
    people = active_people()
    with st.form("new_intervention", clear_on_submit=True):
        c1,c2,c3 = st.columns(3)
        fecha = c1.date_input("Fecha", value=date.today())
        persona = c2.selectbox("Persona", people or ["Persona 1"])
        profesional = c3.text_input("Profesional")
        area = st.text_input("Área")
        objetivo = st.text_input("Objetivo PIA")
        actuacion = st.text_area("Intervención realizada")
        c1,c2,c3 = st.columns(3)
        tipo = c1.selectbox("Tipo apoyo", ["", "Información/orientación", "Supervisión", "Entrenamiento", "Acompañamiento", "Apoyo directo", "Coordinación"])
        intensidad = c2.selectbox("Intensidad", ["", "Sin apoyo", "Baja", "Media", "Alta", "Muy alta"])
        duracion = c3.number_input("Duración (h)", min_value=0.0, step=0.25)
        resultado = st.text_area("Resultado")
        proxima = st.text_input("Próxima actuación")
        revision = st.selectbox("¿Revisión PIA?", ["No", "Sí"])
        observ = st.text_area("Observaciones")
        submitted = st.form_submit_button("Guardar intervención", type="primary")
        if submitted:
            append_record("Intervenciones", {"Fecha": fecha, "Persona": person_id_for_name(persona), "Profesional": profesional, "Área": area, "Objetivo PIA": objetivo, "Intervención realizada": actuacion, "Tipo apoyo": tipo, "Intensidad": intensidad, "Duración (h)": duracion, "Resultado": resultado, "Próxima actuación": proxima, "Revisión PIA": revision, "Observaciones": observ})
            st.success("Intervención guardada.")
    st.subheader("Histórico")
    show_table(load_table("Intervenciones"), "ints")

elif section == "Incidencias":
    st.subheader("Registrar incidencia")
    people = active_people()
    with st.form("new_incident", clear_on_submit=True):
        c1,c2,c3 = st.columns(3)
        fecha = c1.date_input("Fecha", value=date.today())
        persona = c2.selectbox("Persona / ámbito", people or ["Persona 1"])
        gravedad = c3.selectbox("Gravedad", ["Leve", "Moderada", "Grave", "Muy grave"])
        tipo = st.text_input("Tipo")
        indicador = st.text_input("Indicador afectado")
        descripcion = st.text_area("Descripción objetiva")
        impacto = st.text_area("Impacto")
        medida = st.text_area("Medida inmediata")
        c1,c2,c3 = st.columns(3)
        responsable = c1.text_input("Responsable")
        seguimiento = c2.date_input("Fecha seguimiento", value=date.today())
        estado = c3.selectbox("Estado", ["Abierta", "En seguimiento", "Cerrada"])
        correctora = st.text_area("Medida correctora / preventiva")
        submitted = st.form_submit_button("Guardar incidencia", type="primary")
        if submitted:
            append_record("Incidencias", {"Fecha": fecha, "Persona / ámbito": person_id_for_name(persona), "Tipo": tipo, "Gravedad": gravedad, "Indicador afectado": indicador, "Descripción objetiva": descripcion, "Impacto": impacto, "Medida inmediata": medida, "Responsable": responsable, "Fecha seguimiento": seguimiento, "Medida correctora/preventiva": correctora, "Estado": estado})
            st.success("Incidencia guardada.")
    st.subheader("Histórico")
    show_table(load_table("Incidencias"), "incs")

elif section == "Objetivos PIA":
    st.subheader("Añadir / seguir objetivo PIA")
    people = active_people()
    with st.form("new_goal", clear_on_submit=True):
        c1,c2 = st.columns(2)
        persona = c1.selectbox("Persona", people or ["Persona 1"])
        area = c2.text_input("Área")
        objetivo = st.text_area("Objetivo PIA")
        indicador = st.text_input("Indicador / criterio")
        inicial = st.text_area("Situación inicial")
        apoyos = st.text_area("Apoyos previstos")
        c1,c2,c3 = st.columns(3)
        intensidad = c1.selectbox("Intensidad", ["", "Sin apoyo", "Baja", "Media", "Alta", "Muy alta"])
        frecuencia = c2.text_input("Frecuencia")
        grado = c3.number_input("Grado consecución %", min_value=0, max_value=100, step=5)
        evolucion = st.text_area("Evolución")
        dificultades = st.text_area("Dificultades")
        ajustes = st.text_area("Ajustes / nuevas actuaciones")
        c1,c2 = st.columns(2)
        prox = c1.date_input("Próxima revisión", value=date.today())
        estado = c2.selectbox("Estado", ["Activo", "Alcanzado", "Suspendido", "Reformular"])
        submitted = st.form_submit_button("Guardar objetivo", type="primary")
        if submitted:
            append_record("Objetivos PIA", {"Persona": person_id_for_name(persona), "Área": area, "Objetivo PIA": objetivo, "Indicador / criterio": indicador, "Situación inicial": inicial, "Apoyos previstos": apoyos, "Intensidad": intensidad, "Frecuencia": frecuencia, "Fecha inicio": date.today(), "Fecha seguimiento": date.today(), "Evolución": evolucion, "Grado consecución %": grado, "Dificultades": dificultades, "Ajustes / nuevas actuaciones": ajustes, "Próxima revisión": prox, "Estado": estado})
            st.success("Objetivo PIA guardado.")
    st.subheader("Objetivos registrados")
    show_table(load_table("Objetivos PIA"), "goals")

elif section == "Coordinaciones":
    st.subheader("Registrar coordinación")
    people = active_people()
    with st.form("new_coord", clear_on_submit=True):
        c1,c2,c3 = st.columns(3)
        fecha = c1.date_input("Fecha", value=date.today())
        persona = c2.selectbox("Persona", people or ["Persona 1"])
        tipo = c3.text_input("Tipo coordinación")
        entidad = st.text_input("Entidad / recurso")
        contacto = st.text_input("Profesional / contacto")
        modalidad = st.selectbox("Modalidad", ["", "Presencial", "Telefónica", "Videollamada", "Correo", "Otra"])
        motivo = st.text_area("Motivo / objetivo")
        info = st.text_area("Información relevante")
        acuerdos = st.text_area("Acuerdos")
        c1,c2,c3 = st.columns(3)
        responsable = c1.text_input("Responsable")
        plazo = c2.text_input("Plazo / próxima actuación")
        estado = c3.selectbox("Estado", ["Pendiente", "En curso", "Realizada", "Cerrada"])
        observ = st.text_area("Observaciones")
        submitted = st.form_submit_button("Guardar coordinación", type="primary")
        if submitted:
            append_record("Coordinaciones", {"Fecha": fecha, "Persona": person_id_for_name(persona), "Tipo coordinación": tipo, "Entidad / recurso": entidad, "Profesional / contacto": contacto, "Modalidad": modalidad, "Motivo / objetivo": motivo, "Información relevante": info, "Acuerdos": acuerdos, "Responsable": responsable, "Plazo / próxima actuación": plazo, "Estado": estado, "Observaciones": observ})
            st.success("Coordinación guardada.")
    st.subheader("Histórico")
    show_table(load_table("Coordinaciones"), "coords")

elif section == "Datos / copia":
    st.subheader("Archivo de datos")
    st.write("La app trabaja sobre una copia del Excel original, de modo que el archivo subido permanece intacto.")
    with open(WORK_XLSX, "rb") as f:
        st.download_button("Descargar Excel actualizado", f, file_name="Sistema_FUNDATUL_actualizado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if st.button("Restablecer desde el Excel original"):
        shutil.copy2(SOURCE_XLSX, WORK_XLSX)
        clear_cache()
        st.success("Copia de trabajo restablecida.")
    st.info("En Streamlit Community Cloud el almacenamiento local puede reiniciarse al redeplegar o reiniciar la app. Para uso real multiusuario conviene conectar una base de datos persistente gratuita, por ejemplo Supabase.")
